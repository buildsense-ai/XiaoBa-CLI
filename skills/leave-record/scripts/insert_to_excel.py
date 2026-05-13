"""
将标准化请假 JSON 数据写入考勤登记表 Excel。

输入：standardized_approvals.json
输出：在原考勤表基础上追加请假记录的新 Excel 文件

用法：
  python insert_to_excel.py approvals.json "考勤表.xlsx" --output "考勤表_更新.xlsx"
  python insert_to_excel.py approvals.json "考勤表.xlsx" --dry-run  # 预览模式
"""

import argparse
import json
import os
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# 考勤学年列映射：月份 → Excel 列
MONTH_COL = {
    9: 'C', 10: 'D', 11: 'E', 12: 'F',
    1: 'G', 2: 'H', 3: 'I', 4: 'J',
    5: 'K', 6: 'L', 7: 'M', 8: 'N',
}

# 请假类型 → 合计列映射 + 显示名称
SUMMARY_COL = {
    '事假': 'P',
    '病假': 'Q',
}
OTHER_COL = 'O'
OTHER_KEYWORDS = {'调休', '育儿假', '婚假', '丧假', '产检', '产假', '产假调休', '三八调休', '外出支教', '哺乳假'}

# 合计列均分间距（0.5 天为最小单位）
DAY_STEP = 0.5


def hours_to_days(hours: float) -> float:
    """将小时数（企微审批单位）转换为天数（考勤表单位）。

    ≤4h → 0.5 天
    >4h → hours/8，按 0.5 天步进取整
    """
    if hours <= 0:
        return 0.0
    if hours <= 4:
        return 0.5
    raw_days = hours / 8
    return round(raw_days * 2) / 2


def get_leave_category(leave_type: str) -> str:
    """判断请假类型归属的事假/病假/其他假。"""
    lt = leave_type.strip()
    if lt == '事假':
        return '事假'
    if lt == '病假':
        return '病假'
    return '其他假'


def get_summary_value(cell):
    """读取合计列单元格的值（数字），非数字返回 0。"""
    val = cell.value
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return 0.0


def find_teacher_row(ws, name: str):
    """在考勤表 B 列中精确匹配教师姓名，返回行号（1-based），找不到返回 None。"""
    for row in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value  # B 列
        if cell_val and str(cell_val).strip() == name.strip():
            return row
    return None


def make_cell_text(leave_type: str, days: float) -> str:
    """生成月份单元格文本：'{类型}{天数}天'。0.5 天显示为 0.5，整数显示为整数。"""
    if days == int(days):
        day_str = str(int(days))
    else:
        day_str = str(days)
    return f"{leave_type}{day_str}天"


def make_comment_text(start_time: str, leave_type: str, reason: str, hours: float) -> str:
    """生成批注文本，匹配考勤表现有格式。"""
    try:
        dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M")
        day = dt.day
        if dt.hour < 12:
            time_part = "上午"
        else:
            time_part = "下午"
    except ValueError:
        # 回退：只用原始时间字符串
        return f"{start_time} {leave_type}（{reason}）" if reason else f"{start_time} {leave_type}"

    base = f"{day}号{time_part}{leave_type}"
    if reason:
        base += f"（{reason}）"
    return base


def insert_to_excel(approvals: list, excel_path: str, output_path: str, sheet_name: str = None, dry_run: bool = False):
    """将请假记录插入考勤表。"""
    if not os.path.exists(excel_path):
        print(f"错误：考勤表文件不存在 - {excel_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(excel_path)

    # Sheet 选择：优先用 --sheet 参数，否则选名称含"考勤"的最后一个（即年份最旧的）
    target_sheet = None
    if sheet_name:
        for name in wb.sheetnames:
            if sheet_name in name or name in sheet_name:
                target_sheet = name
                break
        if target_sheet is None:
            print(f"错误：未找到匹配 \"{sheet_name}\" 的工作表。可用工作表: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
    else:
        attendance_sheets = [n for n in wb.sheetnames if '考勤' in n]
        if attendance_sheets:
            target_sheet = attendance_sheets[-1]  # 最后一个通常是较早年份
        else:
            target_sheet = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[target_sheet]
    log = []

    for entry in approvals:
        teacher = entry.get("teacher_name", "").strip()
        leave_type = entry.get("leave_type", "").strip()
        start_time = entry.get("start_time", "")
        hours = float(entry.get("hours", 0))
        reason = entry.get("reason", "").strip()
        approval_no = entry.get("approval_no", "")

        if not teacher or not leave_type or hours <= 0:
            log.append({"status": "skip", "entry": entry, "error": "缺少必填字段"})
            continue

        # 1. 姓名匹配
        row = find_teacher_row(ws, teacher)
        if row is None:
            log.append({"status": "error", "teacher": teacher, "error": f"在考勤表中未找到 {teacher}"})
            continue

        # 2. 月份定位
        try:
            dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M")
            month = dt.month
        except ValueError:
            log.append({"status": "error", "teacher": teacher, "error": f"无法解析时间 {start_time}"})
            continue

        col_letter = MONTH_COL.get(month)
        if col_letter is None:
            log.append({"status": "error", "teacher": teacher, "error": f"月份 {month} 超出范围（9月~8月）"})
            continue

        # 3. 计算天数
        days = hours_to_days(hours)
        if days == 0:
            log.append({"status": "skip", "teacher": teacher, "reason": "计算天数为 0"})
            continue

        # 4. 定位合计列
        category = get_leave_category(leave_type)
        if category == '事假':
            summary_col = 'P'
        elif category == '病假':
            summary_col = 'Q'
        else:
            summary_col = 'O'

        # 5. 写入月份单元格
        cell = ws[f"{col_letter}{row}"]
        cell_text = make_cell_text(leave_type, days)

        existing = cell.value
        if existing and str(existing).strip():
            cell.value = f"{existing}；{cell_text}"
        else:
            cell.value = cell_text

        # 6. 批注
        comment_text = make_comment_text(start_time, leave_type, reason, hours)
        if cell.comment:
            cell.comment.text += f"\n{comment_text}"
        else:
            cell.comment = Comment(comment_text, "Administrator")

        # 7. 更新合计列
        summary_cell = ws[f"{summary_col}{row}"]
        current_total = get_summary_value(summary_cell)
        new_total = current_total + days
        if new_total == int(new_total):
            summary_cell.value = int(new_total)
        else:
            summary_cell.value = new_total

        log.append({
            "status": "ok",
            "teacher": teacher,
            "row": row,
            "month": month,
            "col": col_letter,
            "days": days,
            "leave_type": leave_type,
            "category": category,
            "approval_no": approval_no,
        })

    # 输出
    if dry_run:
        print("=== 预览模式，未实际写入 ===\n")
        for entry in log:
            print(json.dumps(entry, ensure_ascii=False))
    else:
        wb.save(output_path)
        print(f"已保存至 {output_path}")
        print(f"共处理 {len(approvals)} 条，成功 {sum(1 for l in log if l['status'] == 'ok')} 条")
        errors = [l for l in log if l['status'] == 'error']
        if errors:
            print(f"失败 {len(errors)} 条：")
            for e in errors:
                print(f"  - {e['teacher']}: {e['error']}")

    return log


def main():
    parser = argparse.ArgumentParser(description="将请假记录插入考勤登记表")
    parser.add_argument("approvals_json", help="标准化请假 JSON 文件路径")
    parser.add_argument("excel_path", help="考勤表 Excel 文件路径")
    parser.add_argument("--output", "-o", default=None, help="输出文件路径（默认在原文件名后加 _更新）")
    parser.add_argument("--sheet", "-s", default=None, help="工作表名称（支持部分匹配，如 \"2024\"）")
    parser.add_argument("--dry-run", action="store_true", help="预览模式，不实际修改文件")
    args = parser.parse_args()

    if not os.path.exists(args.approvals_json):
        print(f"错误：JSON 文件不存在 - {args.approvals_json}", file=sys.stderr)
        sys.exit(1)

    with open(args.approvals_json, 'r', encoding='utf-8') as f:
        approvals = json.load(f)

    output_path = args.output
    if not output_path:
        base = os.path.splitext(args.excel_path)[0]
        output_path = f"{base}_更新.xlsx"

    insert_to_excel(approvals, args.excel_path, output_path, args.sheet, args.dry_run)


if __name__ == "__main__":
    main()
