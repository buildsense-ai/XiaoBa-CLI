"""管理 leave-record 的本地配置文件。"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

CONFIG_PATH = Path.home() / ".config" / "xiaoba" / "leave_record_config.json"


def detect_sheet(excel_path: str) -> str:
    """自动探测考勤表的工作表名。"""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    attendance_sheets = [n for n in wb.sheetnames if '考勤' in n]

    if not attendance_sheets:
        return wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    # 优先选含"2024"的，否则选年份最旧的（最后一个）
    for sheet in attendance_sheets:
        if '2024' in sheet:
            return sheet
    return attendance_sheets[-1]


def cmd_save(excel_path: str, sheet_name: str = None):
    """保存配置。"""
    if not os.path.exists(excel_path):
        print(f"错误：文件不存在 - {excel_path}", file=sys.stderr)
        sys.exit(1)

    if sheet_name is None:
        sheet_name = detect_sheet(excel_path)

    # 验证工作表名
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        # 尝试部分匹配
        matched = [n for n in wb.sheetnames if sheet_name in n or n in sheet_name]
        if matched:
            sheet_name = matched[0]
        else:
            print(f"错误：工作表 \"{sheet_name}\" 不存在。可用工作表: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
    wb.close()

    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    config = {
        "excel_path": os.path.abspath(excel_path),
        "sheet_name": sheet_name,
        "updated_at": datetime.now().isoformat(),
    }
    CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已保存配置：")
    print(f"  文件: {config['excel_path']}")
    print(f"  工作表: {config['sheet_name']}")


def cmd_show():
    """查看当前配置。"""
    if not CONFIG_PATH.exists():
        print("尚未配置。")
        sys.exit(1)
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    print(f"文件路径: {config.get('excel_path', '未设置')}")
    print(f"工作表: {config.get('sheet_name', '未设置')}")
    print(f"更新时间: {config.get('updated_at', '未知')}")

    # 验证文件是否还存在
    excel_path = config.get("excel_path", "")
    if excel_path and not os.path.exists(excel_path):
        print(f"\n⚠ 警告：文件已不存在！请重新配置。")


def main():
    parser = argparse.ArgumentParser(description="管理 leave-record 本地配置")
    sub = parser.add_subparsers(dest="command")

    p_save = sub.add_parser("save", help="保存配置")
    p_save.add_argument("--excel", required=True, help="考勤表 Excel 文件路径")
    p_save.add_argument("--sheet", default=None, help="工作表名（可自动探测）")

    sub.add_parser("show", help="查看当前配置")

    args = parser.parse_args()

    if args.command == "save":
        cmd_save(args.excel, args.sheet)
    elif args.command == "show":
        cmd_show()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
