import argparse
import json

from openpyxl import Workbook

from app.cli import compact_result
from app.manual_ops import swap_lessons
from app.repository import JsonTimetableRepository
from app.services.timetable_service import TimetableService
from app.teacher_package_importer import import_teacher_package


def test_import_teacher_package_reads_class_schedule_workbook(tmp_path):
    package_dir = tmp_path / "package"
    final_dir = package_dir / "最终课表"
    final_dir.mkdir(parents=True)
    workbook_path = final_dir / "小学班级课表（对公版）.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "一1"
    sheet.cell(3, 4).value = "星期一"
    sheet.cell(3, 5).value = "星期二"
    sheet.cell(5, 2).value = 1
    sheet.cell(5, 3).value = "8:50-9:30"
    sheet.cell(5, 4).value = "语文\n(黄奇璇)"
    sheet.cell(5, 5).value = "数学\n(朱臣凤)"
    workbook.save(workbook_path)

    state = import_teacher_package(package_dir)

    assert state["class_counts"]["一年级"] == 1
    assert state["imported_schedule"]["classes"]["一年级(1)"]["周一"][0]["subject"] == "语文"
    assert state["imported_schedule"]["classes"]["一年级(1)"]["周一"][0]["teacher"] == "黄奇璇"
    assert state["imported_schedule"]["periods_by_stage"]["小学"][0]["time"] == "8:50-9:30"


def test_service_uses_imported_schedule_as_current_fact(tmp_path):
    repo = JsonTimetableRepository(tmp_path / "timetable.json")
    repo.save(
        {
            "school_scope": "初中",
            "class_counts": {"七年级": 1},
            "teachers": [],
            "rooms": [],
            "courses": [],
            "messages": [],
            "manual_changes": [],
            "imported_schedule": {
                "class_names": ["七年级(1)"],
                "class_stages": {"七年级(1)": "初中"},
                "periods_by_stage": {"初中": [{"number": 1, "label": "第1节", "time": "8:00-8:40"}]},
                "classes": {
                    "七年级(1)": {
                        "周一": [
                            {
                                "subject": "语文",
                                "teacher": "金婷",
                                "room": "本班教室",
                                "note": "资料包导入",
                                "source": "imported",
                                "movable": True,
                            }
                        ],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    }
                },
            },
        }
    )

    service = TimetableService(repo)
    class_payload = service.show_class("初中", "七年级(1)")
    teacher_payload = service.show_teacher("初中", "金婷")

    assert class_payload["ok"] is True
    assert class_payload["week"]["周一"][0]["subject"] == "语文"
    assert teacher_payload["lessons"] == [
        {"day": "周一", "period": 1, "class": "七年级(1)", "subject": "语文", "room": "本班教室"}
    ]


def test_compact_result_marks_failed_schedule_as_not_ok():
    payload = compact_result({"ok": True, "status": "failed", "message": "暂时没有找到满足全部必须条件的课表"})

    assert payload["ok"] is False


def test_swap_rejects_two_empty_cells():
    schedule = {
        "七年级(1)": {
            "周一": [
                {"subject": "自习", "teacher": "", "room": "本班教室", "source": "empty", "movable": False},
                {"subject": "自习", "teacher": "", "room": "本班教室", "source": "empty", "movable": False},
            ]
        }
    }

    result = swap_lessons(schedule, "七年级(1)", "周一", 1, "七年级(1)", "周一", 2)

    assert result["ok"] is False
    assert "空白" in result["message"]


def test_class_and_teacher_queries_only_return_related_conflicts(tmp_path):
    repo = JsonTimetableRepository(tmp_path / "timetable.json")
    repo.save(
        {
            "school_scope": "初中",
            "class_counts": {"七年级": 3},
            "teachers": [],
            "rooms": [],
            "courses": [],
            "messages": [],
            "manual_changes": [],
            "imported_schedule": {
                "class_names": ["七年级(1)", "七年级(2)", "七年级(3)"],
                "class_stages": {"七年级(1)": "初中", "七年级(2)": "初中", "七年级(3)": "初中"},
                "periods_by_stage": {"初中": [{"number": 1, "label": "第1节", "time": "8:00-8:40"}]},
                "classes": {
                    "七年级(1)": {
                        "周一": [{"subject": "数学", "teacher": "王老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                    "七年级(2)": {
                        "周一": [{"subject": "数学", "teacher": "王老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                    "七年级(3)": {
                        "周一": [{"subject": "英语", "teacher": "李老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                },
            },
        }
    )

    service = TimetableService(repo)
    class_payload = service.show_class("初中", "七年级(3)")
    teacher_payload = service.show_teacher("初中", "李老师")
    validate_payload = service.get_state("初中")

    assert len(validate_payload["conflicts"]) == 1
    assert class_payload["conflicts"] == []
    assert teacher_payload["conflicts"] == []


def test_imported_combined_activity_conflict_is_marked_for_review(tmp_path):
    repo = JsonTimetableRepository(tmp_path / "timetable.json")
    repo.save(
        {
            "school_scope": "初中",
            "class_counts": {"七年级": 2},
            "teachers": [],
            "rooms": [],
            "courses": [],
            "messages": [],
            "manual_changes": [],
            "imported_schedule": {
                "class_names": ["七年级(1)", "七年级(2)"],
                "class_stages": {"七年级(1)": "初中", "七年级(2)": "初中"},
                "periods_by_stage": {"初中": [{"number": 6, "label": "第6节", "time": "14:20-15:00"}]},
                "classes": {
                    "七年级(1)": {
                        "周一": [{"subject": "舞蹈 / 心理健康教育", "teacher": "刘老师、庄老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                    "七年级(2)": {
                        "周一": [{"subject": "心理健康教育 / 舞蹈", "teacher": "庄老师、刘老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                },
            },
        }
    )

    conflicts = TimetableService(repo).get_state("初中")["conflicts"]
    conflict = conflicts[0]

    assert len(conflicts) == 1
    assert conflict["severity"] == "review"
    assert conflict["category"] == "疑似分组/合班/活动课"
    assert conflict["review_required"] is True
    assert conflict["classes"] == ["七年级(1)", "七年级(2)"]


def test_compact_result_keeps_stable_json_fields():
    payload = compact_result({"ok": False, "message": "目标位置已有课程"})

    assert payload["ok"] is False
    assert payload["status"] == "failed"
    assert payload["message"] == "目标位置已有课程"
    assert payload["conflicts"] == []
    assert payload["missing_information"] == []
    assert payload["warnings"] == []
    assert payload["next_actions"] == []


def test_rule_add_on_imported_schedule_says_rule_is_recorded_not_resolved(tmp_path):
    repo = JsonTimetableRepository(tmp_path / "timetable.json")
    repo.save(
        {
            "school_scope": "初中",
            "class_counts": {"七年级": 1},
            "teachers": [],
            "rooms": [],
            "courses": [],
            "messages": [],
            "manual_changes": [],
            "imported_schedule": {
                "class_names": ["七年级(1)"],
                "class_stages": {"七年级(1)": "初中"},
                "periods_by_stage": {"初中": [{"number": 1, "label": "第1节", "time": "8:00-8:40"}]},
                "classes": {
                    "七年级(1)": {
                        "周一": [{"subject": "语文", "teacher": "金婷", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    }
                },
            },
        }
    )

    payload = TimetableService(repo).add_rule("初中", "金婷周一第一节不能上课")

    assert payload["ok"] is True
    assert "规则已记录" in payload["message"]
    assert "尚未自动重排" in payload["message"]


def test_review_only_conflicts_get_confirm_first_next_action(tmp_path):
    repo = JsonTimetableRepository(tmp_path / "timetable.json")
    repo.save(
        {
            "school_scope": "初中",
            "class_counts": {"七年级": 2},
            "teachers": [],
            "rooms": [],
            "courses": [],
            "messages": [],
            "manual_changes": [],
            "imported_schedule": {
                "class_names": ["七年级(1)", "七年级(2)"],
                "class_stages": {"七年级(1)": "初中", "七年级(2)": "初中"},
                "periods_by_stage": {"初中": [{"number": 6, "label": "第6节", "time": "14:20-15:00"}]},
                "classes": {
                    "七年级(1)": {
                        "周一": [{"subject": "舞蹈 / 心理健康教育", "teacher": "刘老师、庄老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                    "七年级(2)": {
                        "周一": [{"subject": "心理健康教育 / 舞蹈", "teacher": "庄老师、刘老师", "room": "本班教室", "source": "imported"}],
                        "周二": [],
                        "周三": [],
                        "周四": [],
                        "周五": [],
                    },
                },
            },
        }
    )

    payload = TimetableService(repo).get_state("初中")

    assert "先确认" in payload["next_actions"][0]
