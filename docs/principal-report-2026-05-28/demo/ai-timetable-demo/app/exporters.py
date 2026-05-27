from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont


PDF_FONT = "STSong-Light"
pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT))


def build_excel_export(school: dict, result: dict) -> bytes:
    workbook = Workbook()
    first_sheet = True
    for class_name in result["class_names"]:
        sheet = workbook.active if first_sheet else workbook.create_sheet()
        first_sheet = False
        safe_title = class_name.replace("(", "").replace(")", "")
        sheet.title = safe_title[:31]
        write_class_sheet(sheet, school, result, class_name)

    rules_sheet = workbook.create_sheet("规则说明")
    rules_sheet.append(["规则类型", "内容"])
    for rule in result.get("applied_rules", []):
        rules_sheet.append(["已应用", rule])
    for rule in result.get("ignored_rules", []):
        rules_sheet.append(["待确认", rule])
    rules_sheet.column_dimensions["A"].width = 14
    rules_sheet.column_dimensions["B"].width = 60

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_class_sheet(sheet, school: dict, result: dict, class_name: str) -> None:
    days = result["days"]
    periods = periods_for_class(result, class_name)
    schedule = result["classes"][class_name]
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 1)
    title_cell = sheet.cell(row=1, column=1, value=f"{class_name}课程表")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header = ["节次/星期"] + days
    sheet.append(header)
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6F4EE")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for period_index, period in enumerate(periods):
        row = [f"{period['label']}\n{period['time']}"]
        for day in days:
            cell = schedule[day][period_index]
            row.append(format_lesson(cell))
        sheet.append(row)

    for row in sheet.iter_rows(min_row=3, max_row=2 + len(periods)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.column_dimensions["A"].width = 16
    for column_index in range(2, len(days) + 2):
        sheet.column_dimensions[chr(64 + column_index)].width = 22
    for row_index in range(3, 3 + len(periods)):
        sheet.row_dimensions[row_index].height = 48


def build_csv_export(school: dict, result: dict, class_name: str | None = None) -> bytes:
    class_name = class_name or result["class_names"][0]
    periods = periods_for_class(result, class_name)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([f"{class_name}课程表"])
    writer.writerow(["节次/星期"] + result["days"])
    schedule = result["classes"][class_name]
    for period_index, period in enumerate(periods):
        row = [f"{period['label']} {period['time']}"]
        for day in result["days"]:
            row.append(format_lesson(schedule[day][period_index]).replace("\n", " / "))
        writer.writerow(row)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def build_pdf_export(school: dict, result: dict, class_name: str | None = None) -> bytes:
    class_name = class_name or result["class_names"][0]
    periods = periods_for_class(result, class_name)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = PDF_FONT
    normal_style = styles["BodyText"]
    normal_style.fontName = PDF_FONT
    normal_style.fontSize = 8
    normal_style.leading = 10

    story = [Paragraph(f"{class_name}课程表", title_style), Spacer(1, 5 * mm)]
    table_data = [["节次/星期"] + result["days"]]
    schedule = result["classes"][class_name]
    for period_index, period in enumerate(periods):
        row = [Paragraph(f"{period['label']}<br/>{period['time']}", normal_style)]
        for day in result["days"]:
            row.append(Paragraph(format_lesson(schedule[day][period_index]).replace("\n", "<br/>"), normal_style))
        table_data.append(row)

    table = Table(table_data, repeatRows=1, colWidths=[28 * mm] + [48 * mm] * len(result["days"]))
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6F4EE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1E2724")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9DED8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def format_lesson(cell: dict) -> str:
    subject = cell.get("subject", "")
    teacher = cell.get("teacher", "")
    room = cell.get("room", "")
    details = [subject]
    if teacher:
        details.append(teacher)
    if room:
        details.append(room)
    return "\n".join(details)


def periods_for_class(result: dict, class_name: str) -> list[dict]:
    stage = result.get("class_stages", {}).get(class_name)
    if stage and "periods_by_stage" in result:
        return result["periods_by_stage"][stage]
    return result["periods"]
