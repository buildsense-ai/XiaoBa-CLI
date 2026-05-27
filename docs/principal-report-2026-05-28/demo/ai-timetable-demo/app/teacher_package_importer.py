from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from tempfile import TemporaryDirectory
import re
from zipfile import ZipFile

from openpyxl import load_workbook

from app.manual_ops import empty_cell


CLASS_SHEET_PATTERN = re.compile(r"^([一二三四五六七八九])(\d+)$")
GRADE_NAMES = {
    "一": "一年级",
    "二": "二年级",
    "三": "三年级",
    "四": "四年级",
    "五": "五年级",
    "六": "六年级",
    "七": "七年级",
    "八": "八年级",
    "九": "九年级",
}
GRADE_STAGE = {
    "一年级": "小学",
    "二年级": "小学",
    "三年级": "小学",
    "四年级": "小学",
    "五年级": "小学",
    "六年级": "小学",
    "七年级": "初中",
    "八年级": "初中",
    "九年级": "初中",
}
DAYS = ["周一", "周二", "周三", "周四", "周五"]


def import_teacher_package(package_path: str | Path) -> dict:
    path = Path(package_path)
    if path.suffix.lower() == ".zip":
        with TemporaryDirectory() as temp_dir:
            with ZipFile(path) as archive:
                archive.extractall(temp_dir)
            return _import_from_directory(Path(temp_dir), source_path=str(path))
    return _import_from_directory(path, source_path=str(path))


def _import_from_directory(root: Path, source_path: str) -> dict:
    if not root.exists():
        raise FileNotFoundError(f"没有找到资料路径：{root}")
    workbooks = select_class_schedule_workbooks(root)
    if not workbooks:
        raise FileNotFoundError("资料包里没有找到班级课表 Excel。")

    imported = {
        "class_names": [],
        "class_stages": {},
        "periods_by_stage": {},
        "classes": {},
        "source_files": [str(path) for path in workbooks],
    }
    class_counts: dict[str, int] = defaultdict(int)

    for workbook_path in workbooks:
        read_class_workbook(workbook_path, imported, class_counts)

    return {
        "school_scope": "全部",
        "class_counts": dict(class_counts),
        "teachers": [],
        "rooms": [],
        "courses": [],
        "messages": [],
        "manual_changes": [],
        "imported_schedule": imported,
        "resolved_schedule": None,
        "source_metadata": {"kind": "teacher_package", "path": source_path},
    }


def select_class_schedule_workbooks(root: Path) -> list[Path]:
    all_workbooks = [
        path
        for path in root.rglob("*.xlsx")
        if "_analysis_outputs" not in path.parts and "_converted_xlsx" not in path.parts
    ]
    self_use = [path for path in all_workbooks if "自用版" in path.name and "班级课表" in path.name]
    if self_use:
        return sorted(self_use)
    public = [path for path in all_workbooks if "班级课表" in path.name and "对公版" in path.name]
    if public:
        return sorted(public)
    scored = [(path, count_class_sheets(path)) for path in all_workbooks]
    scored = [(path, count) for path, count in scored if count > 0]
    if not scored:
        return []
    max_count = max(count for _, count in scored)
    if max_count >= 20:
        return [sorted((path for path, count in scored if count == max_count))[0]]
    return sorted((path for path, _ in scored))


def count_class_sheets(workbook_path: Path) -> int:
    try:
        workbook = load_workbook(workbook_path, read_only=True)
    except Exception:
        return 0
    try:
        return sum(1 for sheet_name in workbook.sheetnames if normalize_class_sheet_name(sheet_name))
    finally:
        workbook.close()


def read_class_workbook(workbook_path: Path, imported: dict, class_counts: dict[str, int]) -> None:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    for sheet in workbook.worksheets:
        class_name = normalize_class_sheet_name(sheet.title)
        if not class_name:
            continue
        grade = class_name.split("(")[0]
        stage = GRADE_STAGE[grade]
        class_counts[grade] = max(class_counts[grade], int(class_name.split("(")[1].rstrip(")")))
        if class_name not in imported["class_names"]:
            imported["class_names"].append(class_name)
        imported["class_stages"][class_name] = stage

        week, periods = read_class_sheet(sheet)
        if periods and stage not in imported["periods_by_stage"]:
            imported["periods_by_stage"][stage] = periods
        imported["classes"][class_name] = week


def normalize_class_sheet_name(sheet_name: str) -> str | None:
    match = CLASS_SHEET_PATTERN.match(str(sheet_name).strip())
    if not match:
        return None
    grade_token, index = match.groups()
    return f"{GRADE_NAMES[grade_token]}({int(index)})"


def read_class_sheet(sheet) -> tuple[dict, list[dict]]:
    week = {day: [] for day in DAYS}
    periods = []
    period_number = 0
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True):
        number = parse_period_number(row)
        if not number:
            continue
        period_number += 1
        periods.append({"number": period_number, "label": f"第{period_number}节", "time": parse_time_label(row) or ""})
        for day_index, day in enumerate(DAYS):
            raw_value = row[3 + day_index] if len(row) > 3 + day_index else None
            week[day].append(parse_lesson_cell(raw_value))
    return week, periods


def parse_period_number(row: tuple) -> int | None:
    for value in row[:3]:
        if isinstance(value, int) and 1 <= value <= 12:
            return value
        text = str(value or "").strip()
        if text.isdigit() and 1 <= int(text) <= 12:
            return int(text)
    return None


def parse_time_label(row: tuple) -> str | None:
    for value in row[:3]:
        text = str(value or "").replace("：", ":").strip()
        if "-" in text and any(char.isdigit() for char in text):
            return text
    return None


def parse_lesson_cell(value) -> dict:
    raw = str(value or "").strip()
    if not raw:
        return empty_cell()
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    subjects = []
    teachers = []
    for line in [part for part in raw.split("\n") if clean_part(part)]:
        teacher_only = re.match(r"^[（(]([^()（）]+)[)）]$", line.strip())
        if teacher_only and subjects and looks_like_teacher(teacher_only.group(1)):
            teachers.append(clean_part(teacher_only.group(1)))
            continue
        subject, teacher = parse_lesson_line(line)
        if subject:
            subjects.append(subject)
        if teacher:
            teachers.append(teacher)
    subject = " / ".join(subjects) or clean_subject(raw)
    teacher = "、".join(dict.fromkeys(teachers))
    return {
        "subject": subject,
        "teacher": teacher,
        "room": room_for_subject(subject),
        "note": raw,
        "source": "imported",
        "movable": True,
    }


def clean_subject(raw: str) -> str:
    text = re.sub(r"[（(][^()（）]+[)）]", "", raw)
    return " / ".join(clean_part(part) for part in text.split("\n") if clean_part(part))


def parse_lesson_line(line: str) -> tuple[str, str]:
    match = re.match(r"(.+)[（(]([^()（）]+)[)）]\s*$", line.strip())
    if match and looks_like_teacher(match.group(2)):
        return clean_part(match.group(1)), clean_part(match.group(2))
    return clean_part(line), ""


def looks_like_teacher(value: str) -> bool:
    text = clean_part(value)
    if not text:
        return False
    if re.fullmatch(r"\d+-\d+", text):
        return False
    if text in {"书法"}:
        return False
    return True


def clean_part(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def room_for_subject(subject: str) -> str:
    if "信息" in subject:
        return "机房A"
    if any(keyword in subject for keyword in ["体育", "运动", "军事体育"]):
        return "操场"
    if any(keyword in subject for keyword in ["科学", "生物"]):
        return "实验室"
    return "本班教室"
