from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel

from app.data import build_demo_school
from app.exporters import build_csv_export, build_excel_export, build_pdf_export
from app.repository import JsonTimetableRepository
from app.services.timetable_service import TimetableService, result_for_request


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

app = FastAPI(title="智能排课 Web 演示版")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


class SolveRequest(BaseModel):
    messages: list[str] = []
    class_name: str | None = None
    class_counts: dict[str, int] | None = None
    school_scope: str = "全部"
    teachers: list[dict] | None = None
    rooms: list[dict] | None = None
    courses: list[dict] | None = None


@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html", headers={"Cache-Control": "no-store"})


@app.get("/api/demo-state")
def demo_state():
    school = build_demo_school()
    return {
        "school": school["school"],
        "stages": school["stages"],
        "grade_settings": school["grade_settings"],
        "grade_counts": school["grade_counts"],
        "days": school["days"],
        "periods": school["periods"],
        "periods_by_stage": school["periods_by_stage"],
        "classes": school["classes"],
        "class_stages": school["class_stages"],
        "teachers": school["teachers"],
        "rooms": school["rooms"],
        "courses": summarize_courses(school["courses"]),
        "fixed_events": school["fixed_events"],
        "example_rules": school["example_rules"],
    }


@app.get("/api/repository-demo-state")
def repository_demo_state(school_scope: str = "初中"):
    """从 repository（skill 写入的真实数据）读取数据，供前端页面渲染"""
    service = TimetableService()
    repo_state = service._load_for_scope(school_scope)
    school, _ = service._build_school(repo_state)
    return {
        "school": school["school"],
        "stages": school["stages"],
        "grade_settings": school["grade_settings"],
        "grade_counts": repo_state.get("class_counts") or school["grade_counts"],
        "days": school["days"],
        "periods": school["periods"],
        "periods_by_stage": school["periods_by_stage"],
        "classes": school["classes"],
        "class_stages": school["class_stages"],
        "teachers": school["teachers"],
        "rooms": school["rooms"],
        "courses": summarize_courses(school["courses"]),
        "fixed_events": school["fixed_events"],
        "example_rules": school["example_rules"],
        "source_mode": service._source_mode_for_state(repo_state),
    }


@app.post("/api/solve")
def solve(request: SolveRequest):
    _, result = build_result_for_export(request)
    return result


@app.post("/api/export/excel")
def export_excel(request: SolveRequest):
    school, result = build_result_for_export(request)
    content = build_excel_export(school, result)
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="timetable.xlsx"'},
    )


@app.post("/api/export/csv")
def export_csv(request: SolveRequest):
    school, result = build_result_for_export(request)
    class_name = resolve_export_class(result, request.class_name)
    content = build_csv_export(school, result, class_name)
    return Response(
        content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="timetable.csv"'},
    )


@app.post("/api/export/pdf")
def export_pdf(request: SolveRequest):
    school, result = build_result_for_export(request)
    class_name = resolve_export_class(result, request.class_name)
    content = build_pdf_export(school, result, class_name)
    return Response(
        content,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="timetable.pdf"'},
    )


@app.post("/api/import-preview")
async def import_preview(file: UploadFile = File(...)):
    content = await file.read()
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    previews = []
    for sheet_name in workbook.sheetnames[:3]:
        sheet = workbook[sheet_name]
        rows = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            rows.append(["" if value is None else str(value) for value in row[:8]])
            if row_index >= 8:
                break
        previews.append({"sheet": sheet_name, "rows": rows})
    return {
        "filename": file.filename,
        "message": "已读取文件预览。第一版演示先使用内置数据排课，后续可把这些表头映射为正式数据。",
        "sheets": previews,
    }


def summarize_courses(courses: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    for course in courses:
        key = (course["grade"], course["subject"])
        if key not in grouped:
            grouped[key] = {
                "grade": course["grade"],
                "subject": course["subject"],
                "weekly_hours": course["weekly_hours"],
                "teacher": course["teacher"],
                "room": course["room"],
                "classes": [],
            }
        grouped[key]["classes"].append(course["class"])
    return list(grouped.values())


def build_result_for_export(request: SolveRequest) -> tuple[dict, dict]:
    if uses_repository_state(request):
        service = TimetableService()
        result = service.get_state(request.school_scope)
        school, _ = service._build_school(service._load_for_scope(request.school_scope))
        return school, result
    return result_for_request(
        messages=request.messages,
        class_counts=request.class_counts,
        school_scope=request.school_scope,
        teachers=request.teachers,
        rooms=request.rooms,
        courses=request.courses,
    )


def uses_repository_state(request: SolveRequest) -> bool:
    return (
        not request.messages
        and request.class_counts is None
        and request.teachers is None
        and request.rooms is None
        and request.courses is None
    )


def resolve_export_class(result: dict, class_name: str | None) -> str:
    class_names = result.get("class_names") or []
    if not class_names:
        raise HTTPException(status_code=400, detail="当前没有可导出的班级，请先检查学段和班级设置。")
    if not class_name:
        return class_names[0]
    if class_name not in result.get("classes", {}):
        raise HTTPException(status_code=400, detail=f"没有找到班级：{class_name}。请先刷新课表，并选择当前学段内的班级。")
    return class_name
